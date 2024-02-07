import openpyxl

wb = openpyxl.load_workbook('table1.xlsx') # Открываем файл
sheet = wb.active # Выбираем активный лист

res = {}

for i in range(sheet.max_row):
    res.update({sheet.cell(row=(i+1), column=1).value: sheet.cell(row=(i+1), column=2).value})
sort_res = sorted(res.items(), key=lambda x: -x[1]) # Сортируем по убыванию

tot = sum(v for k, v in sort_res)
sort_sheet = wb.create_sheet('sort_res') # Создаем новый лист

for i, j in sort_res:
    sort_sheet.append([i, j])
sort_sheet.append(['Total:', tot])

wb.save('table1.xlsx')